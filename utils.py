import requests
import pandas as pd
import numpy as np
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from dateutil.relativedelta import relativedelta
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
import subprocess
import openpyxl



# =============================================================================
# API DATA FETCHING AND PROCESSING
# =============================================================================
def fetch_data(HEADERS, pod, obis_code, start_date, end_date):
    """
    Fetch data from the API endpoint for a given metering point and OBIS code.
    """
    API_URL_TEMPLATE = "https://api.leneda.eu/api/metering-points/{}/time-series"
    url = API_URL_TEMPLATE.format(pod)
    params = {"startDateTime": start_date, "endDateTime": end_date, "obisCode": obis_code}
    try:
        response = requests.get(url, headers=HEADERS, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data for {pod}: {e}")
        return None


def process_api_data(HEADERS, pod, start_date, end_date):
    """
    Process API data into a dictionary of DataFrames for each metering point.
    """
    OBIS_CODES = {
                    "measured_active_production": "1-1:2.29.0",
                    "remaining_production_after_sharing": "1-65:2.29.9"
                 }
    df_list = []
    for obis_label, obis_code in OBIS_CODES.items():
        data = fetch_data(HEADERS, pod, obis_code, start_date, end_date)
        if data and 'items' in data:
            df = pd.DataFrame(data['items'])
            df["meteringPointCode"] = data["meteringPointCode"]
            df["obisCode"] = data["obisCode"]
            df["intervalLength"] = data["intervalLength"]
            df["unit"] = data["unit"]
            df_list.append(df.add_suffix(f"_{obis_label}"))
    if df_list:
        data_frame = pd.concat(df_list, axis=1)
        return data_frame



def calculate_monthly_summaries(df_site, autoconsumption, autoconsumption_price):
    """
    Calculate monthly summaries (aggregations) for production and consumption.
    """
    
    # Use primary timestamp from production data
    df_site.index = pd.to_datetime(df_site["startedAt_measured_active_production"])
        
    if autoconsumption=="Autoconsumption":
        df_site.rename(columns={
                                   "value_measured_active_production": "Production",
                                   "value_remaining_production_after_sharing": "Grid Injection",
                               }, inplace=True)
        df_site["Self-Consumption"] = df_site["Production"] - df_site["Grid Injection"]
        

    elif autoconsumption=="Injection":
        df_site["Self-Consumption"] = 0
        df_site["Production"] = df_site["value_measured_active_production"]
        df_site["Grid Injection"] = df_site["Production"]
        df_site["Self-Consumption Ratio"] = 0
    
    monthly = df_site[["Production", "Grid Injection", "Self-Consumption"]].resample("ME").sum() / 4
    monthly["Self-Consumption Ratio"] = monthly["Self-Consumption"] / monthly["Production"]


    monthly["Pre-VAT Bill"] = monthly["Self-Consumption"] * autoconsumption_price
    monthly["VAT Rate"] = df_site[["VAT Rate"]].resample("ME").mean()
    monthly["VAT Amount"] = monthly["Pre-VAT Bill"] * monthly["VAT Rate"]
    monthly["Total Bill"] = monthly["Pre-VAT Bill"] + monthly["VAT Amount"]

    monthly["Price"] = autoconsumption_price

    monthly["Year"] = monthly.index.year.astype(int)
    monthly["Month"] = monthly.index.strftime('%B')
    monthly.index = monthly.index.tz_localize(None)
    
    return monthly



def apply_excel_formatting(output_file, site_name, pod, capacity, is_injection=False):
    """
    Apply conditional formatting and number formatting to an Excel file.
    """
    wb = load_workbook(output_file)
    ws = wb.active

    # Define fill styles
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="gray125")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Determine the column index for 'Monat'
    df_temp = pd.read_excel(output_file)
    col_idx = list(df_temp.columns).index("Monat") + 1 if "Monat" in df_temp.columns else 1

    # Apply conditional formatting starting from row 3
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
        cell_val = row[col_idx - 1].value
        if cell_val is None:
            prev_val = row[col_idx - 2].value if col_idx - 2 >= 0 else None
            fill = green_fill if i == ws.max_row else (blue_fill if prev_val is None else gray_fill)
            for cell in row:
                cell.fill = fill

    # Set column widths uniformly
    for col in list(ascii_uppercase):
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = 21

    # Number formatting rules
    if not is_injection:
        for col in [7, 10]:
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.0%'
                    if cell.value not in ["", None]:
                        try:
                            cell.value = float(cell.value)
                        except Exception:
                            pass
        for col in [4, 5, 6, 12]:
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.00'
                    if cell.value not in ["", None]:
                        try:
                            cell.value = float(round(cell.value, 0))
                        except Exception:
                            pass
        for col in [9, 11]:
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.00'
                    if cell.value not in ["", None]:
                        try:
                            cell.value = float(cell.value)
                        except Exception:
                            pass
        for col in [8]:
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.000'
                    if cell.value not in ["", None]:
                        try:
                            cell.value = float(cell.value)
                        except Exception:
                            pass
    else:
        # Injection-specific formatting
        for col in [4]:
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.00'
                    if cell.value not in ["", None]:
                        try:
                            cell.value = float(round(cell.value, 0))
                        except Exception:
                            pass

    # Insert header with capacity and timestamp
    header_str = f"Anlage {site_name}, POD {pod}, Kapazität {capacity} kWp   -     Stand: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.insert_rows(1, amount=2)
    ws.cell(row=1, column=1, value=header_str)
    ws.title = site_name
    wb.save(output_file)

    

# =============================================================================
# EXCEL PROCESSING FUNCTIONS
# =============================================================================
def aggregate_dataframe(df, agg_rules, group_period):
    """
    Aggregate a DataFrame based on provided aggregation rules and grouping period.
    """
    grouped = df.groupby(df.index.to_period(group_period)).agg(agg_rules)
    # Recalculate relative consumption where applicable
    if "Eigenverbrauch absolut" in grouped.columns and "Produktion" in grouped.columns:
        grouped["Eigenverbrauch relativ"] = grouped["Eigenverbrauch absolut"] / grouped["Produktion"]
    return grouped
    

def process_sheet(site_name, site_info, pod, df_site_month, sheet_type="autoconsumption"):
    """
    Process a given sheet (autoconsumption or injection) from the Excel input.
    """
    capacity = site_info["Capacity [kWp]"]
    
    print(f"Processing sheet for {site_name} POD {pod} as {sheet_type}...")


    cols = [
            "Jahr",
            "Quartal",
            "Monat",
            "Produktion", 
            "Einspeisung"
            ]

    units = [   "",
                "",
                "",
                "kWh",
                "kWh"
            ]
    if sheet_type == "autoconsumption":
        cols = cols + [
                            "Eigenverbrauch absolut",   
                            "Eigenverbrauch relativ",
                            "Tarif HTVA",
                            "Eigenverbrauch HTVA",
                            "TVA-Satz",
                            "TVA",
                            "Eigenverbrauch TTC"
                        ]
        units = units + [ 
                                "kWh",
                                "%",
                                "€",
                                "€",
                                "%",
                                "€",
                                "€",
                                "",
                                "",
                                ""
                            ]


    df_dec = df_site_month.copy()
    df_dec = df_dec[cols]

    # Define aggregation rules
    if sheet_type == "autoconsumption":
        sum_cols = ['Produktion', 'Einspeisung', 'Eigenverbrauch absolut',
                    'Eigenverbrauch HTVA', 'TVA', 'Eigenverbrauch TTC']
        mean_cols = ['Eigenverbrauch relativ', 'Tarif HTVA', 'TVA-Satz']
        agg_q = {"Jahr": "first", 
                 "Quartal": lambda x: x.iloc[0] + " Total", 
                 "Monat": lambda x: np.nan}
        agg_y = {"Jahr": "first", 
                 "Quartal": lambda x: np.nan, 
                 "Monat": lambda x: np.nan}
        agg_t = {"Jahr": lambda x: "Grand Total", 
                 "Quartal": lambda x: np.nan, 
                 "Monat": lambda x: np.nan}
        for col in sum_cols:
            if col in df_dec.columns:
                agg_q[col] = "sum"
                agg_y[col] = "sum"
                agg_t[col] = "sum"
        for col in mean_cols:
            if col in df_dec.columns:
                agg_q[col] = "mean"
                agg_y[col] = "mean"
                agg_t[col] = "mean"
        agg_t["TVA-Satz"] = lambda x: np.nan
    else:
        # Injection only sums "Produktion"
        agg_q = {"Jahr": "first", 
                 "Quartal": lambda x: x.iloc[0] + " Total", 
                 "Monat": lambda x: np.nan, "Produktion": "sum"}
        agg_y = {"Jahr": "first", 
                 "Quartal": lambda x: np.nan, 
                 "Monat": lambda x: np.nan, 
                 "Produktion": "sum"}
        agg_t = {"Jahr": lambda x: "Grand Total", 
                 "Quartal": lambda x: np.nan, 
                 "Monat": lambda x: np.nan, 
                 "Produktion": "sum"}


    # Perform aggregations
    df_subtotal_q = aggregate_dataframe(df_dec, agg_q, "Q")
    df_subtotal_y = aggregate_dataframe(df_dec, agg_y, "Y")
    df_dec["dummy"] = 1
    df_subtotal_t = df_dec.groupby("dummy").agg(agg_t)
    df_dec.drop(columns=["dummy"], inplace=True)

    if sheet_type == "autoconsumption":
        df_dec["Monat_num"] = df_dec["Monat"]
        df_dec["Monat"] = df_dec["Monat"].dt.year.astype(str) + " - " + df_dec["Monat"].dt.month_name()
    else:
        df_dec["Monat_num"] = df_dec["Monat"]
        df_dec["Monat"] = df_dec["Monat"].dt.strftime("%Y - %B")

    # Concatenate unit row and aggregated data
    df_new = pd.concat([df_dec, df_subtotal_q, df_subtotal_y, df_subtotal_t], axis=0, ignore_index=True)
    df_new = df_new.sort_values(['Jahr', "Quartal", "Monat_num"])
    df_new.drop(columns=["Monat_num"], inplace=True)
    df_units = pd.DataFrame({cols[i]: units[i] for i in range(len(cols))}, index=[0])
    df_new = pd.concat([df_units, df_new], ignore_index=True)
    df_new.reset_index(drop=True, inplace=True)

    # Write output to Excel
    output_file = f"1_Monthly_summaries\Monthly_summary_{site_name}_{pod}.xlsx"
    df_new.to_excel(output_file, index=False)

    # Apply Excel formatting once file is saved
    apply_excel_formatting(output_file, 
                           site_name,
                           pod, 
                           capacity, 
                           is_injection=(sheet_type=="injection"))
    print(f"... done.")


def format_monthly_data(df):
    df_ = df.copy(deep=True)

    df_.rename(columns=
                  {"Production": "Produktion",
                   "Grid Injection": "Einspeisung",
                   "Self-Consumption": "Eigenverbrauch absolut",
                   "Self-Consumption Ratio": 'Eigenverbrauch relativ',
                   "Pre-VAT Bill": 'Eigenverbrauch HTVA',
                   "VAT Amount": 'TVA',
                   "VAT Rate": "TVA-Satz",
                   "Total Bill": 'Eigenverbrauch TTC',
                   "Price": "Tarif HTVA",
                   "Year": "Jahr",
                   "Month": "Monat"},
                  inplace=True)
    
    df_["Jahr"] = df_.index.year.astype(float)
    df_["Quartal"] = pd.to_datetime(df_.index).to_period('Q').astype(str).str[-2:]
    #df["Monat"] = pd.to_datetime(df.index).year.astype(str) + " - " + pd.to_datetime(df.index).month_name()
    df_.index = df_.index.tz_localize(None)
    df_.index.map(lambda x : x.replace(day=1))
    df_["Monat"] = df_.index
    
    df_ind = df_.copy(deep=True)
    df_ind.index = df_ind.index.year.astype(str) + " - " + df_ind.index.month.astype(str)

    return df_





# =============================================================================
# INVOICE GENERATION (PDF)
# =============================================================================
def generate_invoice_for_site(key_pod, df, idx):
    """
    Generate a PDF invoice for a given metering point.
    """
    invoice_filename = f"facture_{key_pod}.pdf"
    doc = SimpleDocTemplate(invoice_filename, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    header_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Add logo if exists
    LOGO_PATH = r"2_Invoices\ecud_logo.png"
    try:
        logo = Image(LOGO_PATH, width=50, height=50)
        elements.append(logo)
    except Exception as e:
        print(f"Logo not found: {e}")
    
    elements.append(Paragraph("Facture", header_style))
    elements.append(Spacer(1, 12))
    
    # Prepare invoice details
    invoice_date = datetime.now().strftime("%d/%m/%Y")
    current_month = datetime.now().month
    invoice_number = f"2025-{current_month:02d}-{idx}"
    vat_number = "LU32018754"
    sender_address = ("Enercoop Uelzechtdall<br/>No. RCS B241591<br/>126 Route de Fischbach<br/>"
                      "L-7447 Lintgen, Luxembourg")
    sender_contact = "Email: clnilles@pt.lu | Tél: +352 691 512 687"
    recipient = "Madame Dany Conter<br/>9, Rue Paul Eyschen<br/>L-7317 Steinsel, Luxembourg"

    amount_due = df['Pre-VAT Bill'].sum() if 'Pre-VAT Bill' in df.columns else 0
    vat_due = df['VAT Amount'].sum() if 'VAT Amount' in df.columns else 0
    total_due = df['Total Bill'].sum() if 'Total Bill' in df.columns else 0
    bank_account = "IBAN: LU56 0019 5655 7212 6000"
    payment_deadline = (datetime.now() + relativedelta(months=1)).strftime("%d/%m/%Y")
    
    invoice_data = [
        ["Date:", invoice_date],
        ["Numéro de la facture:", invoice_number],
        ["Référence:", key_pod],
        ["Période de facturation:", f"{df.index[0].strftime('%d/%m/%Y')} - {df.index[-1].strftime('%d/%m/%Y')}"],
        ["Numéro de TVA:", vat_number],
        ["Adresse:", Paragraph(sender_address, normal_style)],
        ["Contact:", sender_contact],
        ["Destinataire:", Paragraph(recipient, normal_style)],
        ["Total hors TVA:", f"{amount_due:.2f} EUR"],
        ["Taux de TVA:", "8 %"],
        ["Montant total de TVA:", f"{vat_due:.2f} EUR"],
        ["Somme totale à payer:", f"{total_due:.2f} EUR"],
        ["Compte bancaire:", bank_account],
        ["Échéance:", payment_deadline],
    ]
    
    invoice_table = Table(invoice_data, colWidths=[120, 350])
    invoice_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    elements.append(invoice_table)
    doc.build(elements)
    print(f"Invoice generated: {invoice_filename}")

